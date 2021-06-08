import openpyxl

Dict = {}
book = openpyxl.load_workbook("C:\\Users\\srila\\Downloads\\ExcelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=3) #or
print(sheet['A4'].value)
#print(cell.value)
sheet.cell(row=3, column=4).value = "Rewrite"
book.save('ExcelDemo.xlsx')
print(sheet.max_row)
print(sheet.max_column)
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == 'TestCase2':
        for j in range(2, sheet.max_column+1):
            # print(sheet.cell(row=i, column=j).value) #to print values in column 2 from the excel.
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)





#for k in range(1, sheet.max_column+1):
    #print(sheet.cell(row=2, column=k).value)
