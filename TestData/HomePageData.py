import openpyxl
class HomePageData:

    test_HomePage_data = [{"name": "Sri", "email": "test@gmail.com", "gender": "Female"}, {"name": "Keisha", "email": "Wills", "gender": "Male"}]

# importing data from the excel sheet.
# @staticmethod removes the need to declare self in method parameters

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\srila\\Downloads\\ExcelDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value) #to print values in column 2 from the excel.
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]